import math
import random
import pandas as pd
import numpy as np
from deap import base, creator, tools
from collections import defaultdict
from IPython.display import display, HTML
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook

# Pour garantir la reproductibilité
RANDOM_STATE = 42
random.seed(RANDOM_STATE)
np.random.seed(RANDOM_STATE)

########################################
### PARAMÈTRES DE PLANIFICATION      ###
########################################

num_weeks = 10    # Exemple : 10 semaines
num_days = 5      # Du lundi au vendredi
num_slots = 7     # 7 créneaux horaires par jour

########################################
### DÉFINITIONS DES ENTITÉS           ###
########################################

teachers = {
    idx + 1: name
    for idx, name in enumerate([
        "Abdelkrim LAHLOU", "Jihane MALI", "Hugo SANCHEZ", "Christophe RODRIGUES",
        "Zeinab MHANNA", "Frédéric FAUBERTEAU", "Youssef MAZLOUM",
        "Herbert GROSCOT", "Mohamed SOUILAH", "Tristant PINCEAUX", "Maher REBAI",
        "Walter Peretti", "Imen OULED DLALA", "Aline ELLUL", "Guillaume GUERARD", 
        "Sophie DEPEYRE", "Nancy CHENDEB", "Safouane CHENDEB", "Bérengère BRANCHET", 
        "Nédra MELLOULI", "Killian FOURNIER",
        "GRANGEAU Laurent", "Fakir Zachary", "ZHANG Yiru", "GAALOUL Walid", 
        "Eric BERNHART", "Matthéo BODIN", "Constantin TESTU"
    ])
}

modules = {
    idx + 1: name
    for idx, name in enumerate([
        "Infrastructure Technology", "Network", "NoSQL", "Machine Learning",
        "Virtualization and Cloud", "Natural Language Processing",
        "Network Security", "Cybersecurity", "DevOps", "Containerization with Docker",
        "Computational Modeling", "Graph Databases", "Applied Cryptography",
        "NoSQL Query Optimization", "Forensic", "Image Recognition", "Datascience",
        "PTS", "Mobile Devices", "Mathematiques", "Python",
        "ML OPS", "Web dataming", "maths for deep learning", "SOFT SKILLS"
    ])
}

rooms = {
    idx + 1: f"Salle_{101 + idx}"
    for idx in range(20)
}

groups = {
    1: "TD A",
    2: "TD B",
    3: "TD C"
}

# Pour chaque enseignant, on précise quelles matières il peut enseigner.
ENSEIGNANTS_MATIERES = {
    "Abdelkrim LAHLOU": ["Infrastructure Technology", "Network"],
    "Jihane MALI": ["NoSQL", "Machine Learning"],
    "Hugo SANCHEZ": ["Infrastructure Technology", "Virtualization and Cloud"],
    "Christophe RODRIGUES": ["Machine Learning", "Natural Language Processing"],
    "Zeinab MHANNA": ["Network", "Network Security", "Cybersecurity"],
    "Frédéric FAUBERTEAU": ["DevOps", "Containerization with Docker"],
    "Youssef MAZLOUM": ["Computational Modeling", "Graph Databases"],
    "Herbert GROSCOT": ["Applied Cryptography", "NoSQL Query Optimization"],
    "Mohamed SOUILAH": ["Network", "Cybersecurity"],
    "Tristant PINCEAUX": ["Forensic", "Image Recognition"],
    "Maher REBAI": ["Datascience", "PTS"],
    "Walter Peretti": ["Mobile Devices", "Mathematiques"],
    "Imen OULED DLALA": ["Infrastructure Technology", "Datascience"],
    "Aline ELLUL": ["Network Security", "Cybersecurity"],
    "Guillaume GUERARD": ["DevOps", "Virtualization and Cloud"],
    "Sophie DEPEYRE": ["Graph Databases", "Mathematiques"],
    "Nancy CHENDEB": ["NoSQL", "Containerization with Docker"],
    "Safouane CHENDEB": ["Computational Modeling", "Applied Cryptography"],
    "Bérengère BRANCHET": ["Natural Language Processing", "Forensic"],
    "Nédra MELLOULI": ["Machine Learning", "PTS"],
    "Killian FOURNIER": ["Python", "Mathematiques"],
    "GRANGEAU Laurent": ["ML OPS"],
    "Fakir Zachary": ["ML OPS"],
    "ZHANG Yiru": ["Web dataming", "maths for deep learning"],
    "GAALOUL Walid": ["Web dataming", "maths for deep learning"],
    "Eric BERNHART": ["SOFT SKILLS"],
    "Matthéo BODIN": ["SOFT SKILLS"],
    "Constantin TESTU": ["Machine Learning"]
}

teacher_availability = {
    teacher_name: {day: list(range(1, num_slots+1)) for day in range(1, num_days+1)}
    for teacher_name in teachers.values()
}

# Pour un emploi du temps plus chargé, nous augmentons le nombre de cours requis pour chaque groupe.
# Les clés sont les modules obligatoires (ici, ceux définis dans required_counts)
required_counts = {
    group_id: {
        "Machine Learning": 20,
        "Infrastructure Technology": 14,
        "Network": 12,
        "ML OPS": 12,
        "Web dataming": 16,
        "maths for deep learning": 12,
        "SOFT SKILLS": 6,
        "NoSQL": 16,
        "DevOps" : 8,
        "Cybersecurity":8,
        "Containerization with Docker":12,
        "PTS" : 6
        


    }
    for group_id in groups.keys()
}

heures_des_creneaux = {
    1: "08h15 - 09h45",
    2: "10h00 - 11h30",
    3: "11h45 - 13h15",
    4: "13h30 - 15h00",
    5: "15h15 - 16h45",
    6: "17h00 - 18h30",
    7: "18h45 - 20h15"
}

########################################
### FONCTIONS D'ENCODAGE / DÉCODAGE    ###
########################################

def adjust_id(original_id):
    if original_id < 10:
        return original_id
    first_digit = original_id // 10
    return original_id + first_digit

def encode_number(number):
    adjusted_number = adjust_id(number)
    units = adjusted_number % 10
    tens = (adjusted_number // 10) % 10
    hundreds = (adjusted_number // 100) % 10
    if units == 0:
        units = 1
    return f"{units}{tens}{hundreds}"

def decode_number(encoded_str):
    units = int(encoded_str[0])
    tens = int(encoded_str[1])
    hundreds = int(encoded_str[2])
    adjusted_number = units + (tens * 10) + (hundreds * 100)
    if adjusted_number < 10:
        return adjusted_number
    first_digit = adjusted_number // 10
    return adjusted_number - first_digit

def encode(teacher_id, group_id, module_id, room_id, slot_id):
    teacher_code = encode_number(teacher_id)
    group_code = encode_number(group_id)
    module_code = encode_number(module_id)
    room_code = encode_number(room_id)
    slot_code = encode_number(slot_id)
    return f"{teacher_code} {group_code} {module_code} {room_code} {slot_code}"

def decode(encoded_str):
    parts = encoded_str.split()
    if len(parts) == 6:
        mode = "Présentiel" if parts[-1] == "P" else "Distanciel"
        core_parts = parts[:5]
    else:
        mode = "Présentiel"
        core_parts = parts
    teacher_id = decode_number(core_parts[0])
    group_id = decode_number(core_parts[1])
    module_id = decode_number(core_parts[2])
    room_id = decode_number(core_parts[3])
    slot_id = decode_number(core_parts[4])
    teacher_name = teachers.get(teacher_id, f"Professeur {teacher_id}")
    module_name = modules.get(module_id, f"Module {module_id}")
    room_name = rooms.get(room_id, f"Salle_{room_id + 100}") if mode == "Présentiel" else "N/A"
    slot_time = heures_des_creneaux.get(slot_id, "Créneau inconnu")
    return {
        "ID du professeur": teacher_id,
        "Nom du professeur": teacher_name,
        "ID du groupe": group_id,
        "Nom du groupe": groups.get(group_id, f"Groupe {group_id}"),
        "ID Module": module_id,
        "Nom du module": module_name,
        "ID Salle": room_id,
        "Numéro de salle": room_name,
        "ID Créneau": slot_id,
        "Plage horaire": slot_time,
        "Mode": mode
    }

def afficher_recap_encodage():
    recap_data = {
        "Partie encodée": ["Professeur", "Groupe", "Module", "Salle", "Créneau"],
        "Format attendu": [
            "XXX (unit, dizaine, centaine avec décalage)",
            "XXX (unit, dizaine, centaine avec décalage)",
            "XXX (unit, dizaine, centaine avec décalage)",
            "XXX (unit, dizaine, centaine avec décalage)",
            "XXX (unit, dizaine, centaine avec décalage)"
        ],
        "Exemple encodé": [
            "120 (ex: ID 19)",
            "110 (ex: ID 1)",
            "210 (ex: ID 1 ou 20)",
            "130 (ex: ID 3)",
            f"110 (ex: Créneau 1 : {heures_des_creneaux.get(1)})"
        ]
    }
    print("{:<15} {:<50} {:<20}".format("Partie encodée", "Format attendu", "Exemple encodé"))
    print("-" * 90)
    for partie, format_attendu, exemple in zip(
        recap_data["Partie encodée"],
        recap_data["Format attendu"],
        recap_data["Exemple encodé"]
    ):
        print("{:<15} {:<50} {:<20}".format(partie, format_attendu, exemple))

afficher_recap_encodage()

########################################
### INITIALISATION DE L'ALGORITHME     ###
########################################

def create_gene():
    semaine = random.randint(1, num_weeks)
    jour = random.randint(1, num_days)
    module_id = random.choice(list(modules.keys()))
    module_name = modules[module_id]
    possible_teacher_ids = [tid for tid, tname in teachers.items() if module_name in ENSEIGNANTS_MATIERES.get(tname, [])]
    if not possible_teacher_ids:
        raise ValueError(f"Aucun enseignant ne peut enseigner le module {module_name}.")
    teacher_id = random.choice(possible_teacher_ids)
    group_id = random.choice(list(groups.keys()))
    mode = random.choice(["Présentiel", "Distanciel"])
    room_id = random.choice(list(rooms.keys())) if mode == "Présentiel" else 0
    slot_id = random.randint(1, num_slots)
    encoded_core = encode(teacher_id, group_id, module_id, room_id, slot_id)
    encoded = encoded_core + " " + ("P" if mode == "Présentiel" else "D")
    return (semaine, jour, encoded)

# Le nombre total de gènes est déterminé par la somme des cours requis
total_courses = sum(sum(count for count in module_counts.values()) for module_counts in required_counts.values())

# Création conditionnelle des classes DEAP pour éviter les avertissements
if not hasattr(creator, "FitnessMin"):
    creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
if not hasattr(creator, "Individual"):
    creator.create("Individual", list, fitness=creator.FitnessMin)

toolbox = base.Toolbox()
toolbox.register("individual", tools.initRepeat, creator.Individual, create_gene, n=total_courses)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

########################################
### FONCTION DE RÉPARATION            ###
########################################

def repair_individual(ind):
    """Répare les conflits de créneau en réassignant aléatoirement le jour et le slot pour les gènes en conflit."""
    max_iterations = 10
    for _ in range(max_iterations):
        conflict_found = False
        slot_info = defaultdict(list)
        decoded_cache = {}
        for i, gene in enumerate(ind):
            semaine, jour, encoded = gene
            if i not in decoded_cache:
                decoded_cache[i] = decode(encoded)
            key = (semaine, jour, decoded_cache[i]["ID Créneau"])
            slot_info[key].append((i, decoded_cache[i]))
        for key, infos in slot_info.items():
            teacher_counts = defaultdict(int)
            group_counts = defaultdict(int)
            room_counts = defaultdict(int)
            for i, dec in infos:
                teacher_counts[dec["ID du professeur"]] += 1
                group_counts[dec["ID du groupe"]] += 1
                if dec["Mode"] == "Présentiel":
                    room_counts[dec["ID Salle"]] += 1
            for i, dec in infos:
                if (teacher_counts[dec["ID du professeur"]] > 1 or 
                    group_counts[dec["ID du groupe"]] > 1 or 
                    (dec["Mode"] == "Présentiel" and room_counts[dec["ID Salle"]] > 1)):
                    conflict_found = True
                    new_jour = random.randint(1, num_days)
                    new_slot = random.randint(1, num_slots)
                    semaine_val = ind[i][0]
                    teacher_id = dec["ID du professeur"]
                    group_id = dec["ID du groupe"]
                    module_id = dec["ID Module"]
                    room_id = dec["ID Salle"] if dec["Mode"] == "Présentiel" else 0
                    mode = dec["Mode"]
                    new_encoded_core = encode(teacher_id, group_id, module_id, room_id, new_slot)
                    new_encoded = new_encoded_core + " " + ("P" if mode == "Présentiel" else "D")
                    ind[i] = (semaine_val, new_jour, new_encoded)
        if not conflict_found:
            break
    return ind

def repair_missing_courses(ind):
    """
    Répare les cours manquants pour chaque groupe en forçant les gènes
    de l'individu à correspondre aux exigences de required_counts.
    """
    # Construction des comptes actuels et recensement des indices par groupe
    actual_counts = {group_id: {module: 0 for module in required_counts[group_id]} for group_id in required_counts}
    group_gene_indices = {group_id: [] for group_id in required_counts}
    for i, gene in enumerate(ind):
        semaine, jour, encoded = gene
        dec = decode(encoded)
        group_id = dec["ID du groupe"]
        module_name = dec["Nom du module"]
        if group_id in required_counts and module_name in required_counts[group_id]:
            actual_counts[group_id][module_name] += 1
        if group_id in group_gene_indices:
            group_gene_indices[group_id].append(i)
    # Pour chaque groupe et chaque module requis, corriger les manques
    for group_id, mod_req in required_counts.items():
        for module, req in mod_req.items():
            missing = req - actual_counts[group_id][module]
            while missing > 0:
                # Chercher un gène en excès dans ce groupe
                candidate = None
                for i in group_gene_indices[group_id]:
                    semaine, jour, encoded = ind[i]
                    dec = decode(encoded)
                    m = dec["Nom du module"]
                    if m in required_counts[group_id] and actual_counts[group_id][m] > required_counts[group_id][m]:
                        candidate = i
                        break
                if candidate is None:
                    # Sinon, on choisit un gène au hasard dans le groupe
                    candidate = random.choice(group_gene_indices[group_id])
                # Créer un nouveau gène pour combler le manque
                semaine, jour, _ = ind[candidate]
                # Sélectionner un enseignant pouvant enseigner le module manquant
                teacher_candidates = [tid for tid, tname in teachers.items() if module in ENSEIGNANTS_MATIERES.get(tname, [])]
                teacher_id = random.choice(teacher_candidates) if teacher_candidates else 1
                # Récupérer l'ID du module à partir du nom (en parcourant le dictionnaire modules)
                module_id_new = next((mid for mid, mname in modules.items() if mname == module), 1)
                new_room = random.choice(list(rooms.keys()))
                new_slot = random.randint(1, num_slots)
                new_encoded_core = encode(teacher_id, group_id, module_id_new, new_room, new_slot)
                # On choisit aléatoirement le mode
                mode = random.choice(["Présentiel", "Distanciel"])
                new_encoded = new_encoded_core + " " + ("P" if mode == "Présentiel" else "D")
                ind[candidate] = (semaine, jour, new_encoded)
                actual_counts[group_id][module] += 1
                missing -= 1
    return ind

########################################
### FONCTION D'ÉVALUATION             ###
########################################

def evaluate(individual, verbose=False):
    penalties = 0
    actual_counts = {group_id: {module: 0 for module in required_counts[group_id]} for group_id in required_counts}
    teacher_occ = defaultdict(list)
    group_occ = defaultdict(list)
    room_occ = defaultdict(list)
    for gene in individual:
        semaine, jour, encoded = gene
        dec = decode(encoded)
        teacher_id = dec["ID du professeur"]
        group_id = dec["ID du groupe"]
        module_name = dec["Nom du module"]
        room_id = dec["ID Salle"]
        slot_id = dec["ID Créneau"]
        if module_name not in ENSEIGNANTS_MATIERES.get(dec["Nom du professeur"], []):
            if verbose:
                print(f"{dec['Nom du professeur']} ne peut pas enseigner {module_name} ! (+1)")
            penalties += 1
        if slot_id not in teacher_availability.get(dec["Nom du professeur"], {}).get(jour, []):
            if verbose:
                print(f"{dec['Nom du professeur']} n'est pas disponible le jour {jour}, slot {slot_id}. (+1)")
            penalties += 1
        key = (semaine, jour, slot_id)
        teacher_occ[key].append(teacher_id)
        if teacher_occ[key].count(teacher_id) > 1:
            if verbose:
                print(f"Conflit pour l'enseignant {dec['Nom du professeur']} en semaine {semaine}, jour {jour}, slot {slot_id}. (+1)")
            penalties += 1
        group_occ[key].append(group_id)
        if group_occ[key].count(group_id) > 1:
            if verbose:
                print(f"Conflit pour le groupe {groups.get(group_id)} en semaine {semaine}, jour {jour}, slot {slot_id}. (+1)")
            penalties += 1
        room_occ[key].append(room_id)
        if dec["Mode"] == "Présentiel" and room_occ[key].count(room_id) > 1:
            if verbose:
                print(f"Conflit pour la salle {rooms.get(room_id)} en semaine {semaine}, jour {jour}, slot {slot_id}. (+1)")
            penalties += 1
        if group_id in required_counts and module_name in required_counts[group_id]:
            actual_counts[group_id][module_name] += 1
    for group_id, mods in required_counts.items():
        for module, req in mods.items():
            actual = actual_counts[group_id][module]
            if actual < req:
                if verbose:
                    print(f"Manque {req - actual} cours de {module} pour le groupe {groups.get(group_id)}. (+{(req - actual) * 3})")
                penalties += (req - actual) * 3
            elif actual > req:
                if verbose:
                    print(f"Excès de {actual - req} cours de {module} pour le groupe {groups.get(group_id)}. (+{(actual - req) * 3})")
                penalties += (actual - req) * 3
    return (penalties,)

toolbox.register("evaluate", evaluate)
toolbox.register("mate", tools.cxTwoPoint)
toolbox.register("mutate", tools.mutShuffleIndexes, indpb=0.1)
toolbox.register("select", tools.selTournament, tournsize=4)

########################################
### ALGORITHME GÉNÉTIQUE MODIFIÉ      ###
########################################

def main_evolution():
    NGEN = 100       # Générations par cycle
    max_cycles = 300 # Nombre maximal de cycles
    pop = toolbox.population(n=200)
    cycle = 0
    solution_found = False
    while cycle < max_cycles:
        elite = tools.selBest(pop, 1)[0]
        for gen in range(NGEN):
            offspring = toolbox.select(pop, len(pop) - 1)
            offspring = list(map(toolbox.clone, offspring))
            offspring.append(elite)  # Élitisme
            for child1, child2 in zip(offspring[::2], offspring[1::2]):
                if random.random() < 0.7:
                    toolbox.mate(child1, child2)
                    del child1.fitness.values
                    del child2.fitness.values
            for mutant in offspring:
                if random.random() < 0.3:
                    toolbox.mutate(mutant)
                    del mutant.fitness.values
            for ind in offspring:
                repair_individual(ind)
                repair_missing_courses(ind)
            invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
            fitnesses = map(toolbox.evaluate, invalid_ind)
            for ind, fit in zip(invalid_ind, fitnesses):
                ind.fitness.values = fit
            pop[:] = offspring
            elite = tools.selBest(pop, 1)[0]
        best_ind = tools.selBest(pop, 1)[0]
        best_penalty = evaluate(best_ind, verbose=True)[0]
        print(f"Cycle {cycle+1} ({(cycle+1)*NGEN} générations) - Pénalité = {best_penalty}")
        if best_penalty == 0:
            solution_found = True
            print("Emploi du temps sans conflit trouvé !")
            break
        cycle += 1
    if not solution_found:
        print("Aucun emploi du temps sans conflit n'a été trouvé après le nombre maximal de cycles.")
    return best_ind

best_ind = main_evolution()

########################################
### AFFICHAGE DE L'EMPLOI DU TEMPS    ###
########################################

def afficher_emploi_du_temps(chromosome):
    schedule_data = []
    for gene in chromosome:
        semaine, jour, encoded = gene
        dec = decode(encoded)
        schedule_data.append({
            "Semaine": semaine,
            "Jour": ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"][jour-1],
            "Créneau": int(dec["ID Créneau"]),
            "Matière": dec["Nom du module"],
            "Enseignant": dec["Nom du professeur"],
            "Groupe": dec["Nom du groupe"],
            "Mode": dec["Mode"],
            "Salle": dec["Numéro de salle"]
        })
    df = pd.DataFrame(schedule_data)
    df = df.sort_values(by=["Semaine", "Jour", "Créneau"])
    print("Emploi du temps généré :")
    display(HTML(df.to_html(index=False)))
    return df

df = afficher_emploi_du_temps(best_ind)

########################################
### EXPORTATION VERS UN FICHIER EXCEL  ###
########################################

def export_to_excel(df, filename="emploi_du_temps.xlsx"):
    color_presentiel = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    color_distanciel = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    wb = Workbook()
    jours_de_la_semaine = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
    for groupe in df["Groupe"].unique():
        ws = wb.create_sheet(title=groupe)
        ws.cell(row=2, column=1, value="Horaires").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        col_offset = 2
        for semaine in sorted(df["Semaine"].unique()):
            ws.cell(row=1, column=col_offset, value=f"Semaine {semaine}").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(jours_de_la_semaine)-1)
            for col, jour in enumerate(jours_de_la_semaine, start=col_offset):
                ws.cell(row=2, column=col, value=jour).font = Font(bold=True)
                ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
            col_offset += len(jours_de_la_semaine) + 1
        row_offset = 3
        for creneau_id, heure in heures_des_creneaux.items():
            ws.cell(row=row_offset, column=1, value=heure).font = Font(bold=True)
            ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal="center", vertical="center")
            col_offset = 2
            for semaine in sorted(df["Semaine"].unique()):
                for jour_index, jour in enumerate(jours_de_la_semaine):
                    cours = df[(df["Groupe"] == groupe) & (df["Semaine"] == semaine) &
                               (df["Jour"] == jour) & (df["Créneau"] == creneau_id)]
                    if not cours.empty:
                        contenu = "\n".join(
                            f"{row['Matière']}<br>{row['Enseignant']}<br>{row['Groupe']}<br>{row['Mode']}<br>{row['Salle']}"
                            for _, row in cours.iterrows()
                        )
                        mode = cours.iloc[0]["Mode"]
                        fill_color = color_presentiel if mode=="Présentiel" else color_distanciel
                        cell = ws.cell(row=row_offset, column=col_offset+jour_index, value=contenu)
                        cell.fill = fill_color
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                col_offset += len(jours_de_la_semaine) + 1
            row_offset += 1
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(filename)
    print(f"Emploi du temps exporté dans le fichier : {filename}")

export_to_excel(df)
